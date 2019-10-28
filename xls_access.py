import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import os


def workbook_access(filename):
    filename_path = os.path.join("Python Tutorial Supplementary Materials", filename)
    #print(filename_path)

    wb = xl.load_workbook(filename_path)
    sheet = wb['Sheet1']
    #cell = sheet.cell(1, 1)
    #print(cell.value)
    #print(sheet.max_row)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        corrected_price = cell.value * 0.9
        sheet.cell(row, 4).value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename_path)



filenames = os.listdir("Python Tutorial Supplementary Materials")
print(filenames)
workbook_access(filenames[2])